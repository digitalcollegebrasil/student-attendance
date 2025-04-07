import os
import time
import shutil
import pandas as pd
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from datetime import datetime, timedelta
import re
from zeep import Client
import openpyxl
from datetime import datetime, timedelta
import pandas as pd
import os

load_dotenv()

head_office = os.getenv("HEAD_OFFICE")
email_address = os.getenv("SPONTE_EMAIL")
password_value = os.getenv("SPONTE_PASSWORD")

current_dir = os.path.dirname(__file__)
download_dir = os.path.join(current_dir, 'downloads')
base_target_dir = os.path.join(current_dir, 'target')

if not os.path.exists(download_dir):
    os.makedirs(download_dir)
if not os.path.exists(base_target_dir):
    os.makedirs(base_target_dir)

def remove_value_attribute(driver, element):
    driver.execute_script("arguments[0].removeAttribute('value')", element)

def set_input_value(driver, element, value):
    driver.execute_script("arguments[0].value = arguments[1]", element, value)

def get_day_of_week(date):
    return date.strftime("%A")

def move_downloaded_file(download_dir, target_dir, current_date):
    filename = f"Relatorio_{current_date.strftime('%d_%m_%Y')}.xls"
    target_path = os.path.join(target_dir, filename)
    downloaded_files = [f for f in os.listdir(download_dir) if f.endswith('.xls')]
    if downloaded_files:
        latest_file = max(downloaded_files, key=lambda f: os.path.getctime(os.path.join(download_dir, f)))
        shutil.move(os.path.join(download_dir, latest_file), target_path)
        print(f"Moved XLS for {current_date.strftime('%d/%m/%Y')} to {target_path}")

chrome_options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "plugins.always_open_pdf_externally": True
}
chrome_options.add_experimental_option("prefs", prefs)
chrome_options.add_argument("--start-maximized")

start_date_range = input("Enter the start date (dd/mm/yyyy): ")
end_date_range = input("Enter the end date (dd/mm/yyyy): ")
start_date_range = datetime.strptime(start_date_range, "%d/%m/%Y")
end_date_range = datetime.strptime(end_date_range, "%d/%m/%Y")

current_date = start_date_range

def click_element(driver, element):
    driver.execute_script("arguments[0].scrollIntoView();", element)
    driver.execute_script("arguments[0].click();", element)

combined_data = []

while current_date <= end_date_range:
    success = False
    while not success:
        try:
            day_of_week = get_day_of_week(current_date)

            if day_of_week == "Sunday":
                current_date += timedelta(days=1)
                continue

            print(f"Processing date: {current_date.strftime('%d/%m/%Y')} - {day_of_week}")

            driver = webdriver.Chrome(options=chrome_options)
            driver.get("https://www.sponteeducacional.net.br/SPRel/Didatico/Turmas.aspx")
            
            email = driver.find_element(By.ID, "txtLogin")
            email.send_keys(email_address)
            password = driver.find_element(By.ID, "txtSenha")
            password.send_keys(password_value)

            login_button = driver.find_element(By.ID, "btnok")
            login_button.click()
            time.sleep(5)

            print(head_office)
            enterprise = driver.find_element(By.ID, "ctl00_ctl00_spnNomeEmpresa").get_attribute("innerText").strip().replace(" ", "")
            print(enterprise)
            
            combinacoes = {
                ("Aldeota", "DIGITALCOLLEGESUL-74070"): (1, "Acessando a sede Aldeota."),
                ("Aldeota", "DIGITALCOLLEGEBEZERRADEMENEZES-488365"): (1, "Acessando a sede Aldeota."),
                ("Sul", "DIGITALCOLLEGEALDEOTA-72546"): (3, "Acessando a sede Sul."),
                ("Sul", "DIGITALCOLLEGEBEZERRADEMENEZES-488365"): (3, "Acessando a sede Sul."),
                ("Bezerra", "DIGITALCOLLEGEALDEOTA-72546"): (4, "Acessando a sede Bezerra."),
                ("Bezerra", "DIGITALCOLLEGESUL-74070"): (4, "Acessando a sede Bezerra."),
                ("Aldeota", "DIGITALCOLLEGEALDEOTA-72546"): (None, "O script já está na Aldeota."),
                ("Sul", "DIGITALCOLLEGESUL-74070"): (None, "O script já está no Sul."),
                ("Bezerra", "DIGITALCOLLEGEBEZERRADEMENEZES-488365"): (None, "O script já está na Bezerra."),
            }

            resultado = combinacoes.get((head_office, enterprise), (None, "Ação não realizada: combinação não reconhecida."))
            val, message = resultado

            print(message)

            # if val is not None:
            #     driver.execute_script(f"$('#ctl00_hdnEmpresa').val({val});javascript:__doPostBack('ctl00$lnkChange','');")
            #     time.sleep(3)

            if head_office == "Aldeota":
                empresas_button = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_liEmpresas")
                empresas_button.click()
                time.sleep(2)
                ul_element = driver.find_element(By.CSS_SELECTOR, 'ul.nav.nav-pills')
                li_elements = ul_element.find_elements(By.TAG_NAME, 'li')
                if li_elements:
                    first_li = li_elements[0]
                    first_li.click()
                else:
                    print("Nenhum elemento <li> encontrado.")
                time.sleep(2)
            elif head_office == "Sul":
                empresas_button = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_liEmpresas")
                empresas_button.click()
                time.sleep(2)
                aldeota_checkbox = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_cblEmpresas_0")
                aldeota_checkbox.click()
                time.sleep(3)
                sul_checkbox = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_cblEmpresas_1")
                sul_checkbox.click()
                time.sleep(3)
                ul_element = driver.find_element(By.CSS_SELECTOR, 'ul.nav.nav-pills')
                li_elements = ul_element.find_elements(By.TAG_NAME, 'li')
                if li_elements:
                    first_li = li_elements[0]
                    first_li.click()
                else:
                    print("Nenhum elemento <li> encontrado.")
                time.sleep(2)
            elif head_office == "Bezerra":
                empresas_button = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_liEmpresas")
                empresas_button.click()
                time.sleep(2)
                aldeota_checkbox = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_cblEmpresas_0")
                aldeota_checkbox.click()
                time.sleep(3)
                bezerra_checkbox = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_cblEmpresas_2")
                bezerra_checkbox.click()
                time.sleep(3)
                ul_element = driver.find_element(By.CSS_SELECTOR, 'ul.nav.nav-pills')
                li_elements = ul_element.find_elements(By.TAG_NAME, 'li')
                if li_elements:
                    first_li = li_elements[0]
                    first_li.click()
                else:
                    print("Nenhum elemento <li> encontrado.")
                time.sleep(2)
            
            try:
                status_dropdown = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "select2-ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_cmbSituacaoTurma-container"))
                )
                click_element(driver, status_dropdown)
            except TimeoutException:
                print("Status dropdown not clickable")
                driver.quit()
                continue
            time.sleep(1)

            active_status = driver.find_element(By.XPATH, "//*[text()='Vigente']")
            active_status.click()
            time.sleep(5)

            day_of_week_select = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_divDiaSemana")
            day_of_week_select.click()
            time.sleep(1)

            day_of_week_box = None
            if day_of_week == "Monday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Segunda-Feira']")
            elif day_of_week == "Tuesday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Terça-Feira']")
            elif day_of_week == "Wednesday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Quarta-Feira']")
            elif day_of_week == "Thursday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Quinta-Feira']")
            elif day_of_week == "Friday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Sexta-Feira']")
            elif day_of_week == "Saturday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Sábado']")
            elif day_of_week == "Sunday":
                day_of_week_box = driver.find_element(By.XPATH, "//*[text()='Domingo']")
            
            if day_of_week_box:
                day_of_week_box.click()
            time.sleep(1)

            try:
                quantitative_report_checkbox = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_chkRelatorioQuantitativo"))
                )
                click_element(driver, quantitative_report_checkbox)
            except TimeoutException:
                print("Quantitative report checkbox not clickable")
                driver.quit()
                continue
            time.sleep(1)

            try:
                all_classes_checkbox = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_chkMarcarTurmas"))
                )
                click_element(driver, all_classes_checkbox)
            except TimeoutException:
                print("All classes checkbox not clickable")
                driver.quit()
                continue
            time.sleep(3)

            start_date = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_wcdDataInicioFaltasCons_txtData")
            remove_value_attribute(driver, start_date)
            set_input_value(driver, start_date, current_date.strftime("%d/%m/%Y"))

            end_date = driver.find_element(By.ID, "ctl00_ctl00_ContentPlaceHolder1_ContentPlaceHolder2_tab_tabTurmasRegulares_wcdDataTerminoFaltasCons_txtData")
            remove_value_attribute(driver, end_date)
            set_input_value(driver, end_date, current_date.strftime("%d/%m/%Y"))

            try:
                export_checkbox = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "ctl00_ctl00_ContentPlaceHolder1_chkExportar"))
                )
                click_element(driver, export_checkbox)
            except TimeoutException:
                print("Export checkbox not clickable")
                driver.quit()
                continue
            time.sleep(1)

            select2_span = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "select2-ctl00_ctl00_ContentPlaceHolder1_cmbTipoExportacao-container"))
            )
            select2_span.click()
            time.sleep(1)

            option = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//*[text()='Excel Sem Formatação']"))
            )
            option.click()
            time.sleep(1)

            try:
                generate_button = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((By.ID, "ctl00_ctl00_ContentPlaceHolder1_btnGerar_div"))
                )
                click_element(driver, generate_button)
            except TimeoutException:
                print("Generate button not clickable")
                driver.quit()
                continue
            time.sleep(5)

            try:
                move_downloaded_file(download_dir, base_target_dir, current_date)
            
                xls_file_path = os.path.join(base_target_dir, f"Relatorio_{current_date.strftime('%d_%m_%Y')}.xls")

                data = pd.read_excel(xls_file_path, skiprows=3)

                column_mapping = {
                    'Nome': 'Nome',
                    'Professor': 'Professor',
                    'Vagas': 'Vagas',
                    'Integrantes': 'Integrantes',
                    'Trancados': 'Trancados',
                    'Horário': 'Horario',
                    'Não Frequentes': 'NaoFrequente',
                    'Frequentes': 'Frequente',
                    'Dias da Semana': 'DiasSemana'
                }

                selected_columns = {}
                for desired, real in column_mapping.items():
                    if real in data.columns:
                        selected_columns[desired] = data[real]

                selected_columns_df = pd.DataFrame(selected_columns)
                selected_columns_df['Data'] = current_date.strftime("%d/%m/%Y")
                selected_columns_df['Sede'] = head_office

                combined_data.append(selected_columns_df)
            except Exception as e:
                print(f"Erro ao processar a data {current_date.strftime('%d/%m/%Y')}: {str(e)}")
            finally:
                current_date += timedelta(days=1)

            driver.quit()

            success = True
        except Exception as e:
            print(f"Erro ao processar a data {current_date.strftime('%d/%m/%Y')}: {str(e)}")
            driver.quit()

print("Download process completed.")

final_df = pd.concat(combined_data)
final_output_path = os.path.join(current_dir, 'combined_data.xlsx')
final_df.to_excel(final_output_path, index=False)
print(f"Combined data saved to {final_output_path}")

wsdl = 'https://api.sponteeducacional.net.br/WSAPIEdu.asmx?WSDL'
client = Client(wsdl=wsdl)

credenciais = {
    'Aldeota': {
        'codigo_cliente': '72546',
        'token': 'QZUSqqgsLA63'
    },
    'Sul': {
        'codigo_cliente': '74070',
        'token': 'jVNLW7IIUXOh'
    }
}

dias_semana_pt = {
    'Monday': 'Segunda',
    'Tuesday': 'Terça',
    'Wednesday': 'Quarta',
    'Thursday': 'Quinta',
    'Friday': 'Sexta',
    'Saturday': 'Sábado',
    'Sunday': 'Domingo'
}

def parse_intervalo(intervalo):
    dias_semana_pt = {
        'segunda': 'Monday',
        'terça': 'Tuesday',
        'quarta': 'Wednesday',
        'quinta': 'Thursday',
        'sexta': 'Friday',
        'sábado': 'Saturday',
        'domingo': 'Sunday'
    }
    dias_semana_pt_identificados = list(dias_semana_pt.keys())
    
    intervalo = intervalo.lower().strip()
    
    padrao_intervalo = re.compile(
        r'(\bsegunda\b|\bterça\b|\bquarta\b|\bquinta\b|\bsexta\b|\bsábado\b|\bdomingo\b)\s*(?:a|–|-)\s*(\bsegunda\b|\bterça\b|\bquarta\b|\bquinta\b|\bsexta\b|\bsábado\b|\bdomingo\b)', 
        re.IGNORECASE
    )
    intervalos = padrao_intervalo.findall(intervalo)
    
    intervalos_formatados = []
    
    if intervalos:
        for inicio, fim in intervalos:
            if inicio in dias_semana_pt_identificados and fim in dias_semana_pt_identificados:
                indice_inicio = dias_semana_pt_identificados.index(inicio)
                indice_fim = dias_semana_pt_identificados.index(fim)

                dias_intervalo = dias_semana_pt_identificados[indice_inicio:indice_fim + 1]
                intervalos_formatados.extend(dias_intervalo)

    if not intervalos_formatados:
        padrao_dia = re.compile(
            r'\bsegunda\b|\bterça\b|\bquarta\b|\bquinta\b|\bsexta\b|\bsábado\b|\bdomingo\b', 
            re.IGNORECASE
        )
        dias_encontrados = padrao_dia.findall(intervalo)
        dias_encontrados = sorted(set(dias_encontrados), key=lambda x: dias_semana_pt_identificados.index(x))
        intervalos_formatados = dias_encontrados

    return intervalos_formatados

def formatar_intervalo_dias(inicio, fim):
    dias_semana_en = {
        'Segunda': 'Monday',
        'Terça': 'Tuesday',
        'Quarta': 'Wednesday',
        'Quinta': 'Thursday',
        'Sexta': 'Friday',
        'Sábado': 'Saturday',
        'Domingo': 'Sunday'
    }
    dias_semana_pt = {v: k for k, v in dias_semana_en.items()}
    
    dias_semana_pt_identificados = list(dias_semana_pt.keys())
    if inicio in dias_semana_pt_identificados and fim in dias_semana_pt_identificados:
        indice_inicio = dias_semana_pt_identificados.index(inicio)
        indice_fim = dias_semana_pt_identificados.index(fim)
        return dias_semana_pt_identificados[indice_inicio:indice_fim+1]
    return []

def dia_incluso_em_intervalo(dia_semana_pt, intervalo):
    return dia_semana_pt in intervalo

def dia_incluso_em_intervalo_caso_de_ruim(dia_semana_pt, intervalo):
    dias_semana_pt = {
        'Segunda': 'Monday',
        'Terça': 'Tuesday',
        'Quarta': 'Wednesday',
        'Quinta': 'Thursday',
        'Sexta': 'Friday',
        'Sábado': 'Saturday',
        'Domingo': 'Sunday'
    }
    
    dia_en = dias_semana_pt.get(dia_semana_pt, '')
    
    if not dia_en:
        return False

    intervalo = intervalo.lower()
    padrao_dias = re.compile(r'\b(?:segunda|terça|quarta|quinta|sexta|sábado|domingo)\b', re.IGNORECASE)
    dias_encontrados = padrao_dias.findall(intervalo)
    dias_semana_identificados = [dias_semana_pt.get(dia.capitalize(), dia.capitalize()) for dia in dias_encontrados]
    dias_semana_identificados = sorted(set(dias_semana_identificados), key=lambda x: list(dias_semana_pt.values()).index(x))
    
    if not dias_semana_identificados:
        return False

    lista_dias_semana = list(dias_semana_pt.values())
    indice_dia_en = lista_dias_semana.index(dia_en)
    indice_inicio = lista_dias_semana.index(dias_semana_identificados[0])
    indice_fim = lista_dias_semana.index(dias_semana_identificados[-1])

    return indice_inicio <= indice_dia_en <= indice_fim

def formatar_horario(horario):
    horario = horario.lower().strip()
    horario = re.sub(r'\s+', ' ', horario).strip()
    
    intervalo_pattern = re.compile(
        r'(\d{1,2})(?:h|:)?(\d{0,2})?\s*(?:às|a|-\s*das)\s*(\d{1,2})(?:h|:)?(\d{0,2})?', 
        re.IGNORECASE
    )
    intervalos = intervalo_pattern.findall(horario)
    
    horarios_formatados = []
    for inicio_h, inicio_m, fim_h, fim_m in intervalos:
        inicio = f"{inicio_h.zfill(2)}:{inicio_m.zfill(2)}" if inicio_h or inicio_m else ''
        fim = f"{fim_h.zfill(2)}:{fim_m.zfill(2)}" if fim_h or fim_m else ''
        if fim:
            horarios_formatados.append(f"{inicio} às {fim}")
        else:
            horarios_formatados.append(inicio)
    
    horarios_formatados = list(dict.fromkeys(horarios_formatados))
    
    return '; '.join(horarios_formatados)

def formatar_dias_semana(intervalo):
    dias_semana_abreviados = {
        'segunda': 'SEG',
        'terça': 'TER',
        'quarta': 'QUA',
        'quinta': 'QUI',
        'sexta': 'SEX',
        'sábado': 'SAB',
        'domingo': 'DOM'
    }
    
    dias_semana = parse_intervalo(intervalo)
    dias_semana_abrev = [dias_semana_abreviados[dia] for dia in dias_semana]
    return ', '.join(dias_semana_abrev)

def get_turmas_vigentes(data_referencia_dt, dia_semana_referencia_pt, sede):
    codigo_cliente = credenciais[sede]['codigo_cliente']
    token = credenciais[sede]['token']

    try:
        response = client.service.GetTurmas(nCodigoCliente=codigo_cliente, sToken=token, sParametrosBusca='Nome= ;')
        turmas_vigentes = []

        for turma in response:
            data_inicio = datetime.strptime(turma.DataInicio, '%d/%m/%Y') if turma.DataInicio else None
            data_termino = datetime.strptime(turma.DataTermino, '%d/%m/%Y') if turma.DataTermino else None
            
            if turma.Situacao == 'Vigente' and turma.Nome != 'Aulas diversas' and turma.Nome != 'Aulas diversas 2' and turma.Nome != 'AULAS DIVERSAS GT':
                if data_inicio and data_termino and data_inicio <= data_referencia_dt <= data_termino:
                    if dia_incluso_em_intervalo(dia_semana_referencia_pt, turma.Horario) or dia_incluso_em_intervalo_caso_de_ruim(dia_semana_referencia_pt, turma.Horario):
                        turmas_vigentes.append(turma)
        return turmas_vigentes

    except Exception as e:
        print(f"Erro ao obter turmas: {e}")
        return []

def verificar_trancados_turma(turma_id, sede):
    codigo_cliente = credenciais[sede]['codigo_cliente']
    token = credenciais[sede]['token']

    try:
        parametros_busca = f"Situacao=5;TurmaID={turma_id};"
        response = client.service.GetMatriculas(nCodigoCliente=codigo_cliente, sToken=token, sParametrosBusca=parametros_busca)
        total_trancados = 0
        
        if response:
            for matricula in response:
                if matricula.Situacao == "Trancado":
                    total_trancados += 1
        
        return total_trancados
    
    except Exception as e:
        print(f"Erro ao buscar matrículas trancadas para a turma {turma_id}: {e}")
        return 0

def get_frequencia_turma(turma_id, parametros_busca, sede):
    codigo_cliente = credenciais[sede]['codigo_cliente']
    token = credenciais[sede]['token']

    try:
        response = client.service.GetFrequenciaTurma(
            nCodigoCliente=codigo_cliente,
            sToken=token,
            nTurmaID=turma_id,
            sParametrosBusca=parametros_busca
        )
        frequencias = response if isinstance(response, list) else []
        return frequencias
    
    except Exception as e:
        print(f"Erro ao buscar frequência para a turma {turma_id}: {e}")
        return []

def get_diario_aulas(turma_id, data_referencia_dt, disciplina_id, modulo, sede):
    codigo_cliente = credenciais[sede]['codigo_cliente']
    token = credenciais[sede]['token']

    try:
        response = client.service.GetDiarioAulas(
            nCodigoCliente=codigo_cliente,
            sToken=token,
            nTurmaID=turma_id,
            nDisciplinaID=disciplina_id,
            dDataInicio=data_referencia_dt.strftime('%Y-%m-%d'),
            dDataTermino=data_referencia_dt.strftime('%Y-%m-%d'),
            nModulo=modulo
        )

        print(f"Parâmetros de busca para a turma {turma_id}: {disciplina_id}, {modulo}, {data_referencia_dt}, {sede}")
        print(f"Diário de aulas encontrados para a turma {turma_id}: {response}")

        if isinstance(response, list):
            if response:
                for diario in response:
                    retorno = diario['RetornoOperacao']

                    if retorno.startswith("43"):
                        print(f"⚠️ Nenhum registro foi encontrado para a turma {turma_id} ({disciplina_id}, {modulo}, {sede}) na data {data_referencia_dt}.")
                        return None

                    if retorno.startswith("01"):
                        print(f"✅ Operação realizada com sucesso para a turma {turma_id}.")
                        return response

                    if retorno.startswith("02"):
                        print(f"⚠️ Parâmetros inválidos para a turma {turma_id}.")
                        return 'PARÂMETROS_INVÁLIDOS'

            print(f"⚠️ Nenhum diário de aula encontrado para a turma {turma_id}.")
            return []

        print(f"⚠️ A resposta não é uma lista válida. Tipo de resposta: {type(response)}")
        return []

    except Exception as e:
        print(f"Erro ao obter aulas do diário para a turma {turma_id}: {e}")
        return []

def get_quadro_horarios(turma_id, sede):
    codigo_cliente = credenciais[sede]['codigo_cliente']
    token = credenciais[sede]['token']
    parametros_busca = f"sTurmaID={turma_id}"

    try:
        response = client.service.GetQuadroHorarios(
            nCodigoCliente=codigo_cliente,
            sToken=token,
            sParametrosBusca=parametros_busca
        )
        return response
    except Exception as e:
        print(f"Erro ao obter quadro: {e}")
        return None

def registrar_problema(nome_turma, data_referencia, motivo):
    arquivo_excel = 'turmas_com_problemas.xlsx'
    novo_registro = pd.DataFrame([[nome_turma, data_referencia, motivo]], columns=['Turma', 'Data Referência', 'Motivo'])

    if os.path.exists(arquivo_excel):
        df_existente = pd.read_excel(arquivo_excel)
        df = pd.concat([df_existente, novo_registro], ignore_index=True)
        df.drop_duplicates(inplace=True)
    else:
        df = novo_registro
    
    df.to_excel(arquivo_excel, index=False)
    print(f"⚠️ Problema registrado para a turma {nome_turma}: {motivo}")

def main():
    wb_geral = openpyxl.Workbook()
    sheet_frequencias = wb_geral.active
    sheet_frequencias.title = "Frequências"
    
    wb_turmas_live = openpyxl.Workbook()
    sheet_turmas_l = wb_turmas_live.active
    sheet_turmas_l.title = "Turmas Live"
    
    wb_diarios = openpyxl.Workbook()
    sheet_diarios = wb_diarios.active
    sheet_diarios.title = "Diário de Aulas"
    
    headers_diario = ['Turma', 'Data', 'Aula', 'Conteúdo', 'Professor', 
                  'DiarioClasseID', 'NumeroAula', 'DataAula', 'HorarioInicial', 
                  'HorarioFinal', 'ProfessorID', 'Situacao']
    sheet_diarios.append(headers_diario)

    headers = [
        'Data', 'Turma', 'Curso', 'Professor', 'Vagas', 'Integrantes', 'Trancados', 'Horario', 'NaoFrequente', 'Frequente', 'DiasSemana', 'Sede'
    ]
    sheet_frequencias.append(headers)
    sheet_turmas_l.append(headers)

    turmas_com_problemas = []

    for sede in ['Aldeota', 'Sul']:
        for data_referencia in datas:
            data_referencia_dt = datetime.strptime(data_referencia, '%d/%m/%Y')
            dia_semana_referencia_en = data_referencia_dt.strftime('%A')
            dia_semana_referencia_pt = dias_semana_pt.get(dia_semana_referencia_en, '')
            turmas_vigentes = get_turmas_vigentes(data_referencia_dt, dia_semana_referencia_pt, sede)
            
            for turma in turmas_vigentes:
                print(type(turma), turma)

                print("Horário:", turma.Horario)
                print("Nome da Turma:", turma.Nome)
                print("Turma:", turma)

                curso = "N/A"

                if isinstance(turma, str):
                    nome_turma = turma.Nome
                else:
                    nome_turma = getattr(turma, 'Nome', 'Nome_Invalido')

                if nome_turma.startswith("CS"):
                    curso = "Cybersecurity"
                elif nome_turma.startswith("FS"):
                    curso = "Full Stack"
                elif nome_turma.startswith("DA"):
                    curso = "Data Analytics"
                elif nome_turma.startswith("MD"):
                    curso = "Marketing Digital"
                elif nome_turma.startswith("PHP"):
                    curso = "PHP com Laravel"
                elif nome_turma.startswith("UX"):
                    curso = "UX UI"
                elif nome_turma.startswith("PY"):
                    curso = "Python para Dados"
                elif nome_turma.startswith("APM"):
                    curso = "Gerente de Projetos Ágeis"

                professor = turma.ProfessorRegente if hasattr(turma, 'ProfessorRegente') else 'N/A'

                if hasattr(turma, 'Horario'):    
                    horario_turma = formatar_horario(turma.Horario) if hasattr(turma, 'Horario') else 'N/A'

                if hasattr(turma, 'Horario'):
                    dias_semana = formatar_dias_semana(turma.Horario)
                else:
                    print(f"⚠️ A turma {nome_turma} não tem o atributo 'Horario'.")
                    dias_semana = 'Não especificado'

                print(f"Processando a turma {turma.Nome}")

                total_alunos_trancados = verificar_trancados_turma(turma.TurmaID, sede)
                parametros_busca = f'Data={data_referencia}'
                frequencias = get_frequencia_turma(turma.TurmaID, parametros_busca, sede)
                quadro_horarios = get_quadro_horarios(turma.TurmaID, sede)

                if quadro_horarios is None:
                    print(f"Não foi possível obter o quadro de horários para a turma {turma.Nome}.")
                    registrar_problema(nome_turma, data_referencia, "Sem quadro de horários")
                    continue
    
                aulas_no_dia = []
                for horario in quadro_horarios:
                    if hasattr(horario, 'DataAula'):
                        try:
                            data_aula_dt = datetime.strptime(horario.DataAula, '%d/%m/%Y').date()
                            if data_aula_dt == data_referencia_dt.date():
                                diario_aula = get_diario_aulas(turma.TurmaID, data_aula_dt, horario.DisciplinaID, horario.Modulo, sede)

                                if diario_aula == 'PARÂMETROS_INVÁLIDOS':
                                    print(f"⚠️ Parâmetros inválidos para a turma {turma.TurmaID}.")
                                    registrar_problema(nome_turma, data_referencia, "Parâmetros inválidos, erro na API do Sponte ao buscar diário")

                                if diario_aula and isinstance(diario_aula, list):
                                    for diario in diario_aula:
                                        if 'AulasDiario' in diario and 'wsAulasDiario' in diario['AulasDiario']:
                                            aulas_diario = diario['AulasDiario']['wsAulasDiario']
                                            
                                            if aulas_diario:
                                                tem_aula_concluida = any(aula['Situacao'].lower() == 'concluída' for aula in aulas_diario)
                                                
                                                if tem_aula_concluida:
                                                    print(f"✅ Pelo menos uma aula da turma {turma.TurmaID} está concluída.")
                                                    aulas_no_dia.append(horario)
                                                else:
                                                    print(f"⚠️ Nenhuma aula da turma {turma.TurmaID} está como concluída.")
                                                    registrar_problema(nome_turma, data_referencia, "Nenhuma aula consta como concluída")
                                            else:
                                                print(f"🔍 Não há registros de aulas no diário para a turma {turma.TurmaID}.")
                                                registrar_problema(nome_turma, data_referencia, "Sem registros de aula no diário")
                                        else:
                                            print(f"❌ Estrutura inesperada na resposta da API para a turma {turma.TurmaID}.")
                                            registrar_problema(nome_turma, data_referencia, "Erro na estrutura do diário")
                        except ValueError as ve:
                            print(f"Erro ao processar DataAula ({horario.DataAula}): {ve}")
                            registrar_problema(nome_turma, data_referencia, f"Erro na data ({horario.DataAula})")
                    else:
                        print(f"Horário sem DataAula para a turma {nome_turma}.")
                        registrar_problema(nome_turma, data_referencia, "Horário sem DataAula")

                total_presencas_turma = 0
                total_faltas_turma = 0
                total_alunos = 0
                
                alunos_presentes_turma = []
                for item in frequencias:
                    disciplinas = item.Disciplinas
                    if disciplinas:
                        frequencias_disciplinas = disciplinas.wsFrequenciaDisciplinas
                        if frequencias_disciplinas:
                            primeira_disciplina = frequencias_disciplinas[0]
                            alunos = primeira_disciplina.Alunos
                            if alunos:
                                total_alunos = len(alunos.wsFrequenciaAluno)
                                for aluno in alunos.wsFrequenciaAluno:
                                    if aluno.TotalPresencas > 0:
                                        total_presencas_turma += 1
                                        alunos_presentes_turma.append(aluno)
                                    else:
                                        total_faltas_turma += 1
                
                if not aulas_no_dia:
                    if diario_aula == 'PARÂMETROS_INVÁLIDOS' and frequencias:
                        print(f"✅ Aula registrada como concluída para a turma {nome_turma}, apesar de não haver aula no dia, mas possuir frequencia.")
                    else:
                        print(f"Sem aula para a turma {nome_turma} na data {data_referencia}.")
                        registrar_problema(nome_turma, data_referencia, "Sem aula no dia")
                        continue

                if total_faltas_turma == 0 and total_presencas_turma == total_alunos:
                    data_referencia_str = data_referencia_dt.strftime('%d-%m-%Y')

                    if isinstance(turma, str):
                        turma_nome = turma
                    else:
                        turma_nome = getattr(turma, 'Nome', 'Nome_Invalido')

                    filename = f"turma_{turma_nome}_{data_referencia_str}_presenca_completa.xlsx"

                    wb_turma = openpyxl.Workbook()
                    sheet_turma = wb_turma.active
                    sheet_turma.title = turma_nome 

                    sheet_turma.append(['Nome', 'ID', 'TotalPresencas', 'TotalFaltas'])
                    for aluno in alunos_presentes_turma:
                        sheet_turma.append([aluno.NomeAluno, aluno.AlunoID, aluno.TotalPresencas, aluno.TotalFaltas])
                    
                    wb_turma.save(filename)

                vagas = turma.MaxAlunos if hasattr(turma, 'MaxAlunos') else 'N/A'
                integrantes = total_alunos
                nao_frequente = total_faltas_turma
                frequente = total_presencas_turma
                                
                data_row = [
                    data_referencia,
                    nome_turma,
                    curso,
                    professor,
                    vagas,
                    integrantes,
                    total_alunos_trancados,
                    horario_turma,
                    nao_frequente,
                    frequente,
                    dias_semana,
                    sede
                ]

                print("Tipo de dado a ser adicionado:", type(data_row))
                for item in data_row:
                    print(type(item), item)
                
                if len(nome_turma) >= 3 and nome_turma[2].upper() == 'L':
                    sheet_turmas_l.append(data_row)
                else:
                    sheet_frequencias.append(data_row)

    wb_geral.save("frequencia_turmas.xlsx")
    wb_turmas_live.save("turmas_live.xlsx")
    wb_diarios.save("diarios_aulas.xlsx")

if __name__ == "__main__":
    main()