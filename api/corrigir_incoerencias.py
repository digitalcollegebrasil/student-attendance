import re
from zeep import Client
import openpyxl
from datetime import datetime, timedelta
import pandas as pd
import os
from dotenv import load_dotenv
from utils import get_turmas_vigentes, verificar_trancados_turma, get_frequencia_turma, get_diario_aulas, get_quadro_horarios, formatar_horario, formatar_dias_semana, registrar_problema, dias_semana_pt

load_dotenv()

wsdl = 'https://api.sponteeducacional.net.br/WSAPIEdu.asmx?WSDL'
client = Client(wsdl=wsdl)

credenciais = {
    'Aldeota': {
        'codigo_cliente': '72546',
        'token': 'QZUSqqgsLA63'
    },
    'Sul': {
        'codigo_cliente': '74070',
        'token': 'jVNLW7IIUXOh'
    }
}

head_office = os.getenv("HEAD_OFFICE")

if head_office not in credenciais:
    raise ValueError(f"Sede '{head_office}' inválida. Escolha entre: {list(credenciais.keys())}")

data_inicio = datetime.strptime("15/03/2025", "%d/%m/%Y")
data_fim = datetime.strptime("23/03/2025", "%d/%m/%Y")

def main():
    wb_combined = openpyxl.load_workbook("combined_data.xlsx")
    sheet_combined = wb_combined.active

    wb_geral = openpyxl.Workbook()
    sheet_frequencias = wb_geral.active
    sheet_frequencias.title = "Frequências"
    
    wb_turmas_live = openpyxl.Workbook()
    sheet_turmas_l = wb_turmas_live.active
    sheet_turmas_l.title = "Turmas Live"
    
    wb_diarios = openpyxl.Workbook()
    sheet_diarios = wb_diarios.active
    sheet_diarios.title = "Diário de Aulas"

    headers_diario = ['Turma', 'Data', 'Aula', 'Conteúdo', 'Professor', 
                      'DiarioClasseID', 'NumeroAula', 'DataAula', 'HorarioInicial', 
                      'HorarioFinal', 'ProfessorID', 'Situacao']
    sheet_diarios.append(headers_diario)

    headers = [
        'Data', 'Turma', 'Curso', 'Professor', 'Vagas', 'Integrantes', 'Trancados', 'Horario', 
        'NaoFrequente', 'Frequente', 'DiasSemana', 'Sede'
    ]
    sheet_frequencias.append(headers)
    sheet_turmas_l.append(headers)

    turmas_com_problemas = []
    data = None

    for row in sheet_combined.iter_rows(min_row=2, values_only=True):
        nome, professor, vagas, integrantes, trancados, horario, nao_frequentes, frequentes, dias_semana, data, sede = row
        data = datetime.strptime(data, "%d/%m/%Y") if isinstance(data, str) else data

        print("Processando turma:", nome)
        print("Data:", data)
        print("Sede:", sede)

        data_referencia_dt = datetime.strptime(str(data.date()), '%Y-%m-%d')
        dia_semana_referencia_en = data_referencia_dt.strftime('%A')
        dia_semana_referencia_pt = dias_semana_pt.get(dia_semana_referencia_en, '')

        turmas_vigentes = get_turmas_vigentes(data_referencia_dt, dia_semana_referencia_pt, sede, credenciais)
        print("Turmas vigentes:", len(turmas_vigentes))

        for turma in turmas_vigentes:
            if turma.Nome == nome:
                curso = "N/A"
                nome_turma = getattr(turma, 'Nome', 'Nome_Invalido')

                if nome_turma.startswith("CS"):
                    curso = "Cybersecurity"
                elif nome_turma.startswith("FS"):
                    curso = "Full Stack"
                elif nome_turma.startswith("DA"):
                    curso = "Data Analytics"
                elif nome_turma.startswith("MD"):
                    curso = "Marketing Digital"
                elif nome_turma.startswith("PHP"):
                    curso = "PHP com Laravel"
                elif nome_turma.startswith("UX"):
                    curso = "UX UI"
                elif nome_turma.startswith("PY"):
                    curso = "Python para Dados"
                elif nome_turma.startswith("APM"):
                    curso = "Gerente de Projetos Ágeis"

                professor = turma.ProfessorRegente if hasattr(turma, 'ProfessorRegente') else 'N/A'
                horario_turma = formatar_horario(turma.Horario) if hasattr(turma, 'Horario') else 'N/A'
                dias_semana = formatar_dias_semana(turma.Horario) if hasattr(turma, 'Horario') else 'Não especificado'

                total_alunos_trancados = verificar_trancados_turma(turma.TurmaID, sede, credenciais)
                parametros_busca = f'Data={data_referencia_dt.date()}'
                frequencias = get_frequencia_turma(turma.TurmaID, parametros_busca, sede, credenciais)
                quadro_horarios = get_quadro_horarios(turma.TurmaID, sede, credenciais)

                if quadro_horarios is None:
                    print(f"Não foi possível obter o quadro de horários para a turma {turma.Nome}.")
                    registrar_problema(nome_turma, data_referencia_dt.date(), "Sem quadro de horários")
                    continue

                aulas_no_dia = []
                for horario in quadro_horarios:
                    if hasattr(horario, 'DataAula'):
                        try:
                            data_aula_dt = datetime.strptime(horario.DataAula, '%d/%m/%Y').date()
                            if data_aula_dt == data_referencia_dt.date():
                                diario_aula = get_diario_aulas(turma.TurmaID, data_aula_dt, horario.DisciplinaID, horario.Modulo, sede, credenciais)
                                if diario_aula == 'PARÂMETROS_INVÁLIDOS':
                                    print(f"⚠️ Parâmetros inválidos para a turma {turma.TurmaID}.")
                                    registrar_problema(nome_turma, data_referencia_dt.date(), "Parâmetros inválidos, erro na API do Sponte ao buscar diário")
                                if diario_aula and isinstance(diario_aula, list):
                                    for diario in diario_aula:
                                        if 'AulasDiario' in diario and 'wsAulasDiario' in diario['AulasDiario']:
                                            aulas_diario = diario['AulasDiario']['wsAulasDiario']
                                            
                                            if aulas_diario:
                                                tem_aula_concluida = any(aula['Situacao'].lower() == 'concluída' for aula in aulas_diario)
                                                
                                                if tem_aula_concluida:
                                                    aulas_no_dia.append(horario)
                                                else:
                                                    registrar_problema(nome_turma, data_referencia_dt.date(), "Nenhuma aula consta como concluída")
                                        else:
                                            registrar_problema(nome_turma, data_referencia_dt.date(), "Erro na estrutura do diário")
                        except ValueError as ve:
                            print(f"Erro ao processar DataAula ({horario.DataAula}): {ve}")
                            registrar_problema(nome_turma, data_referencia_dt.date(), f"Erro na data ({horario.DataAula})")

                total_presencas_turma = 0
                total_faltas_turma = 0
                total_alunos = 0

                alunos_presentes_turma = []
                for item in frequencias:
                    disciplinas = item.Disciplinas
                    if disciplinas:
                        frequencias_disciplinas = disciplinas.wsFrequenciaDisciplinas
                        if frequencias_disciplinas:
                            primeira_disciplina = frequencias_disciplinas[0]
                            alunos = primeira_disciplina.Alunos
                            if alunos:
                                total_alunos = len(alunos.wsFrequenciaAluno)
                                for aluno in alunos.wsFrequenciaAluno:
                                    if aluno.TotalPresencas > 0:
                                        total_presencas_turma += 1
                                        alunos_presentes_turma.append(aluno)
                                    else:
                                        total_faltas_turma += 1

                if total_faltas_turma == 0 and total_presencas_turma == total_alunos:
                    data_referencia_str = data_referencia_dt.strftime('%d-%m-%Y')
                    if isinstance(turma, str):
                        turma_nome = turma
                    else:
                        turma_nome = getattr(turma, 'Nome', 'Nome_Invalido')
                    filename = f"turma_{turma_nome}_{data_referencia_str}_presenca_completa.xlsx"
                    wb_turma = openpyxl.Workbook()
                    sheet_turma = wb_turma.active
                    sheet_turma.title = turma_nome
                    sheet_turma.append(['Nome', 'ID', 'TotalPresencas', 'TotalFaltas'])
                    for aluno in alunos_presentes_turma:
                        sheet_turma.append([aluno.NomeAluno, aluno.AlunoID, aluno.TotalPresencas, aluno.TotalFaltas])
                    
                    wb_turma.save(filename)

                vagas = turma.MaxAlunos if hasattr(turma, 'MaxAlunos') else 'N/A'
                integrantes = total_alunos
                nao_frequente = total_faltas_turma
                frequente = total_presencas_turma

                data_formatada = data_referencia_dt.strftime('%d/%m/%Y')
                                
                data_row = [
                    data_formatada,
                    nome_turma,
                    curso,
                    professor,
                    vagas,
                    integrantes,
                    total_alunos_trancados,
                    horario_turma,
                    nao_frequente,
                    frequente,
                    dias_semana,
                    sede
                ]

                if len(nome_turma) >= 3 and nome_turma[2].upper() == 'L':
                    sheet_turmas_l.append(data_row)
                else:
                    sheet_frequencias.append(data_row)
            else:
                print("Turma não encontrada:", nome)
                turmas_com_problemas.append(nome)

    wb_geral.save("frequencia_turmas.xlsx")
    wb_turmas_live.save("turmas_live.xlsx")
    wb_diarios.save("diarios_aulas.xlsx")

if __name__ == "__main__":
    main()